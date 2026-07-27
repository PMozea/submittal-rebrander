"""
codebook.py - load the KCC/Trane model-number digit definitions out of the two
nomenclature workbooks.

Each sheet lays its digits out as multi-column blocks:  "DIGIT n - TITLE" header,
then rows of  code | description.  A few blocks (heat capacity, motor HP) carry
several description columns (IF | ELECTRIC | HOT WATER, or ECM | Belt | Direct
Drive); those are returned as a list so the caller can pick the right column.
"""
import os
import re

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))

CUR_SHEETS = {"OAB5": "OAB rev 5", "OAN5": "OAN rev5", "REV6": "OAD-OAN rev6"}
HYB_SHEET = "Viking OA"

SEQ = "ABCDEFGHJKLMNPRSTUVWXYZ"   # KCC code ladder: no I, O or Q

_HEAD = re.compile(r"^DIGITS?\s+([\d,\s and&]+?)\s*[-\u2013]\s*(.+)$", re.I)


def _grid(path, sheet):
    ws = load_workbook(path, data_only=True)[sheet]
    g = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                s = str(v).strip().replace("\n", " ")
                if s:
                    g[(r, c)] = s
    return g, ws.max_row, ws.max_column


def _load_sheet(path, sheet, ncols=4):
    g, maxr, maxc = _grid(path, sheet)
    heads = []
    for (r, c), v in g.items():
        m = _HEAD.match(v)
        if m:
            digs = tuple(int(x) for x in re.findall(r"\d+", m.group(1)))
            heads.append((r, c, digs, m.group(2).strip()))
    bycol = {}
    for h in heads:
        bycol.setdefault(h[1], []).append(h)
    out = {}
    for col, hs in bycol.items():
        hs.sort(key=lambda h: h[0])
        for i, (r, c, digs, title) in enumerate(hs):
            stop = hs[i + 1][0] if i + 1 < len(hs) else maxr + 1
            codes = {}
            prev = None
            for rr in range(r + 1, stop):
                code = g.get((rr, c))
                if code is not None and _HEAD.match(code):
                    break
                if code is None:
                    # Some blocks (e.g. OAN rev5 d28) leave the code cell blank on
                    # the last few rows even though the option exists. Continue the
                    # KCC code ladder from the previous row rather than dropping it.
                    if prev is None or prev not in SEQ:
                        continue
                    nxt = SEQ.index(prev) + 1
                    if nxt >= len(SEQ):
                        continue
                    code = SEQ[nxt]
                    if not any(g.get((rr, c + k)) for k in range(1, ncols + 1)):
                        continue
                # The OAN sheet interleaves single-letter cabinet applicability
                # flags (A/K/N) between the code and its description, so gather
                # every non-trivial cell to the right and drop the 1-char flags.
                clean = []
                for k in range(1, ncols + 1):
                    v = g.get((rr, c + k))
                    if v is None:
                        continue
                    if _HEAD.match(v):
                        break
                    if len(v) == 1:          # applicability flag, not a description
                        continue
                    clean.append(v)
                if not clean:
                    continue
                codes.setdefault(code, clean)
                prev = code
            if digs not in out or len(codes) > len(out[digs][1]):
                out[digs] = (title, codes)
    return out


def load_books(cur_path=None, hyb_path=None):
    cur_path = cur_path or os.path.join(HERE, "cur.xlsx")
    hyb_path = hyb_path or os.path.join(HERE, "hyb.xlsx")
    books = {k: _load_sheet(cur_path, s) for k, s in CUR_SHEETS.items()}
    books["HYB"] = _load_sheet(hyb_path, HYB_SHEET)
    return books


# ---------------------------------------------------------------- positions
# Nominal digit positions that are never printed in the model number.
UNUSED = {"REV5": {10, 19}, "REV6": {10, 20, 30, 40, 50, 60}}


def positions(book, rev):
    """Ordered list of (digit-tuple, title) actually present in the string."""
    fields = sorted(book.items(), key=lambda kv: kv[0][0])
    return [(d, t) for d, (t, _) in fields if d[0] not in UNUSED[rev]]


def split_model(model, book, rev):
    """Walk the printed characters onto digit positions. Returns {digit_tuple: code}."""
    s = model.replace("-", "").strip()
    out, i = {}, 0
    for digs, _title in positions(book, rev):
        n = len(digs)
        out[digs] = s[i:i + n]
        i += n
    return out, s[i:]
